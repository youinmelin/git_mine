# -*- coding: UTF-8 -*-
import my_logging
import os 
import openpyxl

logger = my_logging.MyLogging().getlogger()

class SaveListToXlsx():

    def write_xlsx_file(self,*arg):
        '''将传入的列表存入excel表格中
            alldata: 是传入的list数据，要求是一个一维或者二位列表
        '''
        #print('started to write xlsx')
        file_name = 'result.xlsx'
#        logger.debug(arg)
        alldata,file_path,sheet_name = arg
        if file_name in os.listdir(file_path+'.'):
            wb = openpyxl.load_workbook('%s%s'%(file_path,file_name))
        else:
            wb = openpyxl.Workbook()
        try:
            #如果存在同名的表，就删除
            wb.remove(wb[sheet_name])
        except Exception as e:
            logger.debug(e)
        ws = wb.create_sheet(sheet_name)
        row_num = 1
        for t_row in alldata:
            col_num = 1
            for t_col in t_row:
                ws.cell(row=row_num,column=col_num,value=t_col)
                col_num += 1
            row_num += 1
        wb.save('%s%s'%(file_path,file_name))
        #print('xlsx writed')

